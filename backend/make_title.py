import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_title_page(workbook_path: str, bank_name: str, date_start: str, date_end: str, boss_name: str):
    """
    Добавляет титульный лист в существующий Excel файл
    """
    # Загружаем существующий файл
    wb = load_workbook(workbook_path)
    
    # Создаем новый лист и перемещаем его в начало
    if "Титульник" in wb.sheetnames:
        ws = wb["Титульник"]
    else:
        ws = wb.create_sheet("Титульник", 0)
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 5

    # Заголовок
    ws.merge_cells('B4:D4')
    title_cell = ws['B4']
    title_cell.value = "РАБОЧИЕ ДОКУМЕНТЫ АУДИТОРА"
    title_cell.font = Font(name='Times New Roman', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Информация о банке
    ws.merge_cells('B6:D6')
    bank_cell = ws['B6']
    bank_cell.value = f"по аудиту {bank_name}"
    bank_cell.font = Font(name='Times New Roman', size=14)
    bank_cell.alignment = Alignment(horizontal='center')

    # Период проверки
    ws.merge_cells('B8:D8')
    period_cell = ws['B8']
    period_cell.value = f"за период с {date_start} по {date_end}"
    period_cell.font = Font(name='Times New Roman', size=12)
    period_cell.alignment = Alignment(horizontal='center')

    # Руководитель проверки
    ws.merge_cells('B12:D12')
    ws['B11'].value = "Руководитель проверки:"
    ws['B11'].font = Font(name='Times New Roman', size=12)
    boss_cell = ws['B12']
    boss_cell.value = boss_name
    boss_cell.font = Font(name='Times New Roman', size=12, bold=True)
    boss_cell.alignment = Alignment(horizontal='center')

    # Добавляем рамки для основного контента
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws['B4:D12']:
        for cell in row:
            cell.border = thin_border

    # Сохраняем файл
    wb.save(workbook_path)
    return workbook_path 